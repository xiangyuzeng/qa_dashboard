#!/usr/bin/env python3
"""
Excel export — builds a 7-sheet styled workbook FROM SCRATCH with openpyxl.Workbook()
(no template). The old May workbook was the 6-sheet "no-café" variant with zero risk
fills (the "too plain" artifact); this rewrite authors the workbook in code so the
styling — 5-level Risk-Level cell fills, navy bilingual frozen headers, autofilter — is
the source of truth.

Sheets: 1) 月度摘要 Monthly Summary · 2) 食品安全主表 · 3) 进口出口监管 · 4) 州地方法规 ·
5) 咖啡馆检查 (incl. 门店编号) · 6) 数据源日志 · 7) 字段说明.
Output is a STATIC file the dashboard links to (public/exports/monthly_report.xlsx).

Run: npm run prep:export   (python3 prep/export_xlsx.py)  ·  needs: pip install openpyxl
"""
import json
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(ROOT, "data", "v2")
OUT_DIR = os.path.join(ROOT, "public", "exports")
OUT = os.path.join(OUT_DIR, "monthly_report.xlsx")

# ── Column orders — MIRROR src/lib/schema.ts SHEET1..4_COLUMNS (kept in sync; the
#    lengths are asserted here AND printed by prep/validate.ts so any drift surfaces). ──
SHEET1_COLS = ["no", "category", "chineseTitle", "englishTitle", "source", "publicationDate",
               "chineseSummary", "englishSummary", "sourceUrl", "riskLevel", "relevanceNotes",
               "recommendedAction"]
SHEET2_COLS = ["no", "jurisdiction", "regulatoryAgency", "brand", "establishmentType", "storeName",
               "establishmentId", "address", "inspectionDate", "inspectionType", "inspectionResult",
               "score", "grade", "violationCode", "chineseViolationSummary", "englishViolationSummary",
               "violationSeverity", "standardizedCategory", "followupRequired", "sourceType",
               "riskLevel", "sourceUrlOrDocRef", "recommendedAction"]
SHEET3_COLS = ["no", "category", "chineseTitle", "englishTitle", "agency", "countryRegion",
               "productInvolved", "publicationDate", "regulatoryAction", "chineseSummary",
               "englishSummary", "importExportImpact", "documentationRequirement", "riskLevel",
               "sourceUrl", "recommendedAction"]
SHEET4_COLS = ["no", "jurisdiction", "regulationBillName", "chineseTitle", "englishTitle", "status",
               "publicationPassageDate", "effectiveDate", "coveredEntities", "keyRequirements",
               "chineseSummary", "englishSummary", "businessImpact", "riskLevel", "sourceUrl",
               "recommendedAction"]
assert len(SHEET1_COLS) == 12 and len(SHEET2_COLS) == 23 and len(SHEET3_COLS) == 16 and len(SHEET4_COLS) == 16

# ── V2.5 compliance-domain column orders — MIRROR src/lib/schema.ts SHEET5..8_COLUMNS. ──
SHEET5_COLS = ["no", "jurisdiction", "regulationBillName", "chineseTitle", "englishTitle", "agency",
               "applicabilityThreshold", "appliesToUs", "status", "effectiveDate", "keyRequirements",
               "chineseSummary", "englishSummary", "businessImpact", "enforcementRecord", "riskLevel",
               "sourceUrl", "recommendedAction"]
SHEET6_COLS = ["no", "jurisdiction", "codeStandardName", "chineseTitle", "englishTitle", "agency",
               "codeCitation", "status", "effectiveDate", "coveredEntities", "keyRequirements",
               "chineseSummary", "englishSummary", "businessImpact", "inspectionCitationRecord", "penalty",
               "riskLevel", "sourceUrl", "recommendedAction"]
SHEET7_COLS = ["no", "jurisdiction", "regulationName", "chineseTitle", "englishTitle", "agency",
               "applicabilityThreshold", "appliesToUs", "status", "effectiveDate", "keyRequirements",
               "chineseSummary", "englishSummary", "businessImpact", "riskLevel", "sourceUrl",
               "recommendedAction"]
SHEET8_COLS = ["no", "jurisdiction", "regulationName", "chineseTitle", "englishTitle", "agency",
               "applicabilityThreshold", "appliesToUs", "keyRequirements", "complaintEnforcementRecord",
               "status", "effectiveDate", "riskLevel", "sourceUrl", "recommendedAction"]
assert len(SHEET5_COLS) == 18 and len(SHEET6_COLS) == 19 and len(SHEET7_COLS) == 17 and len(SHEET8_COLS) == 15

# ── Bilingual headers (data-key → "中文\nEnglish") ──
HDR1 = {
    "no": "序号\nNo.", "category": "类别\nCategory", "chineseTitle": "中文标题\nTitle (ZH)",
    "englishTitle": "英文标题\nTitle (EN)", "source": "来源\nSource", "publicationDate": "发布日期\nPublished",
    "chineseSummary": "中文摘要\nSummary (ZH)", "englishSummary": "英文摘要\nSummary (EN)",
    "sourceUrl": "原文链接\nSource URL", "riskLevel": "风险等级\nRisk Level",
    "relevanceNotes": "相关性说明\nRelevance", "recommendedAction": "建议行动\nRecommended Action",
}
HDR2 = {
    "no": "序号\nNo.", "jurisdiction": "地区\nJurisdiction", "regulatoryAgency": "监管机构\nAgency",
    "brand": "品牌\nBrand", "establishmentType": "门店类型\nEstablishment Type", "storeName": "门店名称\nStore Name",
    "establishmentId": "门店编号\nEstablishment ID", "address": "地址\nAddress",
    "inspectionDate": "检查日期\nInspection Date", "inspectionType": "检查类型\nType",
    "inspectionResult": "检查结果\nResult", "score": "分数\nScore", "grade": "等级\nGrade",
    "violationCode": "违规代码\nViolation Code", "chineseViolationSummary": "中文违规摘要\nViolation (ZH)",
    "englishViolationSummary": "英文违规摘要\nViolation (EN)", "violationSeverity": "严重程度\nSeverity",
    "standardizedCategory": "标准化类别\nStd. Category", "followupRequired": "是否复查\nFollow-up",
    "sourceType": "数据来源方式\nSource Type", "riskLevel": "风险等级\nRisk Level",
    "sourceUrlOrDocRef": "原文链接/文件\nSource Ref", "recommendedAction": "建议行动\nRecommended Action",
}
HDR3 = {
    "no": "序号\nNo.", "category": "类别\nCategory", "chineseTitle": "中文标题\nTitle (ZH)",
    "englishTitle": "英文标题\nTitle (EN)", "agency": "监管机构\nAgency", "countryRegion": "国家/地区\nCountry/Region",
    "productInvolved": "涉及产品\nProduct", "publicationDate": "发布日期\nPublished",
    "regulatoryAction": "监管动作\nAction", "chineseSummary": "中文摘要\nSummary (ZH)",
    "englishSummary": "英文摘要\nSummary (EN)", "importExportImpact": "进出口影响\nImport/Export Impact",
    "documentationRequirement": "文件要求\nDocumentation", "riskLevel": "风险等级\nRisk Level",
    "sourceUrl": "原文链接\nSource URL", "recommendedAction": "建议行动\nRecommended Action",
}
HDR4 = {
    "no": "序号\nNo.", "jurisdiction": "地区\nJurisdiction", "regulationBillName": "法规/法案名称\nBill / Rule",
    "chineseTitle": "中文标题\nTitle (ZH)", "englishTitle": "英文标题\nTitle (EN)", "status": "状态\nStatus",
    "publicationPassageDate": "发布/通过日期\nPublished/Passed", "effectiveDate": "生效日期\nEffective",
    "coveredEntities": "适用对象\nCovered Entities", "keyRequirements": "核心要求\nKey Requirements",
    "chineseSummary": "中文摘要\nSummary (ZH)", "englishSummary": "英文摘要\nSummary (EN)",
    "businessImpact": "业务影响\nBusiness Impact", "riskLevel": "风险等级\nRisk Level",
    "sourceUrl": "原文链接\nSource URL", "recommendedAction": "建议行动\nRecommended Action",
}

# ── V2.5 compliance-domain headers (one combined map covers SHEET5..8 keys) ──
HDR_DOMAIN = {
    "no": "序号\nNo.", "jurisdiction": "地区\nJurisdiction", "agency": "监管机构\nAgency",
    "regulationBillName": "法规/法案名称\nRule / Bill", "codeStandardName": "法规/标准名称\nCode / Standard",
    "regulationName": "法规名称\nRegulation", "chineseTitle": "中文标题\nTitle (ZH)", "englishTitle": "英文标题\nTitle (EN)",
    "applicabilityThreshold": "适用门槛\nApplicability Threshold", "appliesToUs": "适用我方\nApplies to Us",
    "status": "状态\nStatus", "effectiveDate": "生效日期\nEffective", "keyRequirements": "核心要求\nKey Requirements",
    "chineseSummary": "中文摘要\nSummary (ZH)", "englishSummary": "英文摘要\nSummary (EN)",
    "businessImpact": "业务影响\nBusiness Impact", "enforcementRecord": "执法/投诉记录\nEnforcement",
    "codeCitation": "条款/代码\nCode Citation", "coveredEntities": "适用对象\nCovered Entities",
    "inspectionCitationRecord": "检查/处罚记录\nInspection/Citation", "penalty": "罚款\nPenalty",
    "complaintEnforcementRecord": "投诉/执法记录\nComplaint/Enforcement",
    "riskLevel": "风险等级\nRisk Level", "sourceUrl": "原文链接\nSource URL",
    "recommendedAction": "建议行动\nRecommended Action",
}
HDR5 = HDR6 = HDR7 = HDR8 = HDR_DOMAIN

MODULE_LABEL = {
    "food_safety": "食品安全 Food Safety", "import": "进口出口 Import", "regulation": "州地方法规 Regulation",
    "inspection": "咖啡馆检查 Inspection", "sentiment": "负面舆情 Sentiment",
    "labor": "用工合规 Labor", "building": "建筑与职业安全 Building", "environment": "环境卫生 Environmental",
    "consumer": "消费者与员工保护 Consumer",
}
# mirrors the bilingual glosses in src/lib/schema.ts PullStatusEnum comments
STATUS_LABEL = {
    "fetched": "已抓取 Fetched", "filtered": "已筛选 Filtered", "no_update": "未发现更新 No update",
    "manual": "人工 Manual", "excluded": "排除 Excluded",
}

# ── Styling ──
NAVY = "1F3A5F"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
RISK_LEVELS = ["高风险", "中风险", "低风险", "关注", "信息参考"]
# Light fills aligned with the on-screen badge palette (src/lib/colors.ts RISK_COLORS).
RISK_FILL = {
    "高风险": PatternFill("solid", fgColor="F4CCCC"),   # red   (high  ← #C00000)
    "中风险": PatternFill("solid", fgColor="FCE4D6"),   # amber (medium ← #B45309)
    "低风险": PatternFill("solid", fgColor="E2EFDA"),   # green (low   ← #15803D)
    "信息参考": PatternFill("solid", fgColor="E7EBF0"),  # slate (info  ← #64748B)
    "关注": PatternFill("solid", fgColor="CFE8EF"),     # teal  (watch ← #0891B2, distinct from low's green)
}
HIGH_FONT = Font(color="C00000", bold=True)
HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(color=WHITE, bold=True, size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP_TOP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")

WIDTHS = {
    "no": 6, "riskLevel": 12, "category": 18, "publicationDate": 13, "publicationPassageDate": 14,
    "effectiveDate": 13, "inspectionDate": 13, "score": 8, "grade": 8, "establishmentId": 16,
    "establishmentType": 16, "jurisdiction": 16, "brand": 14, "agency": 18, "regulatoryAgency": 18,
    "countryRegion": 16, "productInvolved": 28, "regulatoryAction": 16, "status": 14,
    "chineseSummary": 46, "englishSummary": 46, "chineseTitle": 30, "englishTitle": 30,
    "chineseViolationSummary": 38, "englishViolationSummary": 38, "recommendedAction": 38,
    "sourceUrl": 36, "sourceUrlOrDocRef": 36, "relevanceNotes": 24, "coveredEntities": 26,
    "keyRequirements": 38, "businessImpact": 34, "importExportImpact": 34, "documentationRequirement": 30,
    "regulationBillName": 26, "storeName": 24, "address": 26, "inspectionType": 16,
    "inspectionResult": 16, "violationCode": 16, "violationSeverity": 14, "standardizedCategory": 22,
    "followupRequired": 12, "sourceType": 18, "source": 18,
    # V2.5 compliance-domain columns
    "codeStandardName": 28, "regulationName": 28, "applicabilityThreshold": 26, "appliesToUs": 14,
    "enforcementRecord": 32, "codeCitation": 20, "inspectionCitationRecord": 30, "penalty": 22,
    "complaintEnforcementRecord": 30, "_default": 18,
}
SUMMARY_COLS = 7  # width of the bespoke summary sheet (A..G)


def load(name):
    with open(os.path.join(DATA, name), encoding="utf-8") as f:
        return json.load(f)


def servable(r):
    return r.get("reviewed") and r.get("reviewStatus") != "rejected"


def fmt(key, val):
    if key == "appliesToUs":
        return "未评估 N/A" if val is None else ("适用 Yes" if val else "暂不适用 No")
    if val is None:
        return ""
    if "date" in key.lower() and isinstance(val, str) and val:
        return val.replace("-", "/")
    return val


# ── shared styling helpers ──
def style_header_row(ws, headers):
    for ci, label in enumerate(headers):
        c = ws.cell(row=1, column=ci + 1, value=label)
        c.fill, c.font, c.alignment, c.border = HEADER_FILL, HEADER_FONT, HEADER_ALIGN, BORDER
    ws.row_dimensions[1].height = 30


def apply_col_widths(ws, columns):
    for ci, key in enumerate(columns):
        ws.column_dimensions[get_column_letter(ci + 1)].width = WIDTHS.get(key, WIDTHS["_default"])


def write_data_row(ws, row_idx, columns, record, seq):
    high = record.get("riskLevel") == "高风险"
    for ci, key in enumerate(columns):
        cell = ws.cell(row=row_idx, column=ci + 1)
        cell.value = seq if key == "no" else fmt(key, record.get(key))
        cell.border, cell.alignment = BORDER, WRAP_TOP
        if high:
            cell.fill, cell.font = RISK_FILL["高风险"], HIGH_FONT
        if key == "riskLevel":
            lvl = record.get("riskLevel")
            if lvl in RISK_FILL:
                cell.fill = RISK_FILL[lvl]
            if lvl == "高风险":
                cell.font = HIGH_FONT


def finalize_table(ws, n_cols, n_rows):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(n_cols)}{max(1, n_rows) + 1}"


def build_table_sheet(wb, title, columns, header_map, records):
    ws = wb.create_sheet(title)
    style_header_row(ws, [header_map[c] for c in columns])
    for i, rec in enumerate(records):
        write_data_row(ws, 2 + i, columns, rec, i + 1)
    apply_col_widths(ws, columns)
    finalize_table(ws, len(columns), len(records))
    return ws


# ── bespoke sheet 1: 月度摘要 ──
def _section(ws, r, label):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=SUMMARY_COLS)
    c = ws.cell(row=r, column=1, value=label)
    c.fill, c.font = HEADER_FILL, Font(color=WHITE, bold=True, size=11)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 20
    return r + 1


def _line(ws, r, zh, en):
    parts = [p for p in [zh, en] if p]
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=SUMMARY_COLS)
    c = ws.cell(row=r, column=1, value=("•  " + "  /  ".join(parts)) if parts else "")
    c.alignment = WRAP_TOP
    return r + 1


def build_summary_sheet(wb, meta, matrix, verdicts=None):
    ws = wb.create_sheet("月度摘要")
    summary = meta.get("summary") or {}
    counts = meta.get("counts", {})
    period = meta.get("reportingPeriod", {})

    # title band
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=SUMMARY_COLS)
    c = ws.cell(row=1, column=1, value=summary.get("reportNameZh", "月度摘要 Monthly Summary"))
    c.fill, c.font, c.alignment = HEADER_FILL, Font(color=WHITE, bold=True, size=14), HEADER_ALIGN
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=SUMMARY_COLS)
    ws.cell(row=2, column=1, value=summary.get("reportNameEn", "")).font = Font(italic=True, color="64748B")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=SUMMARY_COLS)
    ws.cell(row=3, column=1, value=f"数据截止 Data as of: {meta.get('dataAsOf','')}   ·   周期 Period: {period.get('label','')}")
    r = 5

    r = _section(ws, r, "监测范围 Scope")
    r = _line(ws, r, summary.get("scopeZh"), summary.get("scopeEn"))
    r += 1

    r = _section(ws, r, "排除项 Exclusions")
    for e in summary.get("exclusions", []):
        r = _line(ws, r, e.get("zh"), e.get("en"))
    r += 1

    # KPI table
    r = _section(ws, r, "关键指标 KPIs")
    kpi = [("食品安全主表\nFood Safety", counts.get("regulatory", 0)),
           ("进口出口\nImport", counts.get("importExport", 0)),
           ("州地方法规\nRegulation", counts.get("regulation", 0)),
           ("咖啡馆检查\nInspections", counts.get("inspections", 0)),
           ("负面舆情\nSentiment", counts.get("sentiment", 0)),
           ("高风险\nHigh-risk", counts.get("highRisk", 0)),
           ("关注\nWatch", counts.get("watch", 0))]
    for ci, (h, _v) in enumerate(kpi):
        c = ws.cell(row=r, column=ci + 1, value=h)
        c.fill, c.font, c.alignment, c.border = HEADER_FILL, HEADER_FONT, HEADER_ALIGN, BORDER
    ws.row_dimensions[r].height = 28
    r += 1
    for ci, (_h, v) in enumerate(kpi):
        c = ws.cell(row=r, column=ci + 1, value=v)
        c.alignment, c.border = CENTER, BORDER
        if ci == 5:
            c.font = HIGH_FONT
    r += 2

    # risk-mix matrix (module × 5-level)
    r = _section(ws, r, "风险分布 Risk Mix (模块 × 风险等级)")
    c = ws.cell(row=r, column=1, value="模块 \\ 风险\nModule \\ Risk")
    c.fill, c.font, c.alignment, c.border = HEADER_FILL, HEADER_FONT, HEADER_ALIGN, BORDER
    for ci, lvl in enumerate(RISK_LEVELS):
        c = ws.cell(row=r, column=ci + 2, value=lvl)
        c.fill, c.font, c.alignment, c.border = HEADER_FILL, HEADER_FONT, HEADER_ALIGN, BORDER
    c = ws.cell(row=r, column=len(RISK_LEVELS) + 2, value="合计\nTotal")
    c.fill, c.font, c.alignment, c.border = HEADER_FILL, HEADER_FONT, HEADER_ALIGN, BORDER
    ws.row_dimensions[r].height = 24
    r += 1
    for m in ["food_safety", "import", "regulation", "labor", "building", "environment", "consumer", "inspection", "sentiment"]:
        mc = ws.cell(row=r, column=1, value=MODULE_LABEL[m])
        mc.font, mc.border = Font(bold=True), BORDER
        total = 0
        for ci, lvl in enumerate(RISK_LEVELS):
            v = matrix.get(m, {}).get(lvl, 0)
            total += v
            cell = ws.cell(row=r, column=ci + 2, value=v)
            cell.alignment, cell.border = CENTER, BORDER
            if v and lvl in RISK_FILL:
                cell.fill = RISK_FILL[lvl]
        tc = ws.cell(row=r, column=len(RISK_LEVELS) + 2, value=total)
        tc.alignment, tc.font, tc.border = CENTER, Font(bold=True), BORDER
        r += 1
    r += 1

    # 合规姿态 Compliance Posture snapshot — one row per applicability rule (from the engine verdicts).
    if verdicts:
        r = _section(ws, r, "合规姿态 Compliance Posture (适用性 / Applicability)")
        for ci, h in enumerate(["规则 Rule", "我方当前值 Our Value", "阈值 Threshold", "适用? Applies?"]):
            c = ws.cell(row=r, column=ci + 1, value=h)
            c.fill, c.font, c.alignment, c.border = HEADER_FILL, HEADER_FONT, HEADER_ALIGN, BORDER
        r += 1
        for v in verdicts:
            status = v.get("status", "na")
            ws.cell(row=r, column=1, value=v.get("nameZh") or v.get("nameEn", "")).border = BORDER
            our = v.get("ourValue")
            ws.cell(row=r, column=2, value=("待补充 Pending" if our is None and status == "na" else ("—" if our is None else our))).border = BORDER
            ws.cell(row=r, column=3, value=v.get("threshold") if v.get("threshold") is not None else "—").border = BORDER
            vc = ws.cell(row=r, column=4, value=("待核实 To verify" if (status == "na" and v.get("needsVerification")) else APPLIES_LABEL.get(status, status)))
            vc.border = BORDER
            if status in APPLIES_FILL:
                vc.fill = APPLIES_FILL[status]
                if status == "applies":
                    vc.font = HIGH_FONT
            r += 1
        r += 1

    r = _section(ws, r, "关键要点 Key Highlights")
    for h in summary.get("keyHighlights", []):
        r = _line(ws, r, h.get("zh"), h.get("en"))
    r += 1

    # high-risk items table
    r = _section(ws, r, "高风险事项 High-Risk Items")
    for ci, h in enumerate(["模块 Module", "标题 Title", "风险 Risk", "链接 Ref"]):
        c = ws.cell(row=r, column=ci + 1, value=h)
        c.fill, c.font, c.alignment, c.border = HEADER_FILL, HEADER_FONT, HEADER_ALIGN, BORDER
    r += 1
    for it in summary.get("highRiskItems", []):
        ws.cell(row=r, column=1, value=MODULE_LABEL.get(it.get("module"), it.get("module", ""))).border = BORDER
        tc = ws.cell(row=r, column=2, value=it.get("titleZh") or it.get("titleEn") or it.get("recordId", ""))
        tc.alignment, tc.border = WRAP_TOP, BORDER
        rc = ws.cell(row=r, column=3, value=it.get("riskLevel", ""))
        rc.border = BORDER
        lvl = it.get("riskLevel")
        if lvl in RISK_FILL:
            rc.fill = RISK_FILL[lvl]
        if lvl == "高风险":
            rc.font = HIGH_FONT
        ws.cell(row=r, column=4, value=it.get("href", "")).border = BORDER
        r += 1
    r += 1

    r = _section(ws, r, "关键行动 Key Actions")
    for a in summary.get("keyActions", []):
        r = _line(ws, r, a.get("zh"), a.get("en"))

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 40
    for col in ["C", "D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 18
    ws.freeze_panes = "A2"
    return ws


# ── bespoke sheet 6: 数据源日志 ──
def build_sources_sheet(wb, meta):
    ws = wb.create_sheet("数据源日志")
    style_header_row(ws, ["数据源\nSource", "模块\nModule", "状态\nStatus", "记录数\nRecords",
                          "原文链接\nURL", "备注\nNotes"])
    prov = meta.get("provenance", [])
    for i, p in enumerate(prov):
        row = 2 + i
        vals = [p.get("name", ""), MODULE_LABEL.get(p.get("module"), p.get("module", "")),
                STATUS_LABEL.get(p.get("status"), p.get("status", "")), p.get("recordCount", 0),
                p.get("endpointOrUrl") or "", p.get("stalenessNote") or ""]
        for ci, v in enumerate(vals):
            c = ws.cell(row=row, column=ci + 1, value=v)
            c.border, c.alignment = BORDER, WRAP_TOP
    for ci, w in enumerate([34, 22, 18, 10, 50, 40]):
        ws.column_dimensions[get_column_letter(ci + 1)].width = w
    finalize_table(ws, 6, len(prov))
    return ws


# ── bespoke sheet 7: 字段说明 ──
def build_field_guide_sheet(wb):
    ws = wb.create_sheet("字段说明")
    style_header_row(ws, ["工作表 Sheet", "字段 Field", "说明 Description"])
    rows = [
        ("月度摘要 Monthly Summary", "—", "由 meta.summary + counts 生成：报告名称、范围、排除项、KPI、风险分布矩阵、关键要点、高风险事项、关键行动。"),
        ("食品安全主表 / 进口出口监管 / 州地方法规 / 咖啡馆检查", "序号 No.", "行序号（顺序生成）。"),
        ("（通用 Common）", "风险等级 Risk Level", "5 级：高风险 / 中风险 / 低风险 / 关注 / 信息参考；单元格按等级着色。"),
        ("（通用 Common）", "建议行动 Recommended Action", "针对该条目的质量行动建议。"),
        ("咖啡馆检查 Inspections", "门店编号 Establishment ID", "NYC camis / Boston licenseno / 设施编号；重复违规检测的稳定关联键。"),
        ("数据源日志 Sources", "状态 Status", "抓取状态：已抓取 / 已筛选 / 未发现更新 / 人工 / 排除。"),
        ("负面舆情 Sentiment", "（无独立工作表 no own sheet）", "舆情仅汇总于月度摘要（KPI + 风险分布矩阵）；明细见仪表盘 /sentiment，绝不转载文章正文。"),
    ]
    for i, (sheet, field, desc) in enumerate(rows):
        row = 2 + i
        for ci, v in enumerate([sheet, field, desc]):
            c = ws.cell(row=row, column=ci + 1, value=v)
            c.border, c.alignment = BORDER, WRAP_TOP
    r = 2 + len(rows) + 1

    ws.cell(row=r, column=1, value="风险等级图例 Risk Legend").font = Font(bold=True, color=NAVY)
    r += 1
    risk_desc = {
        "高风险": "直接影响 (high — direct impact)", "中风险": "中等影响 (medium)",
        "低风险": "较低影响 (low)", "信息参考": "信息参考 (informational)",
        "关注": "暂无直接影响，但存在复发 / 恶化趋势 / 升级可能 (watch)",
    }
    for lvl in RISK_LEVELS:
        cc = ws.cell(row=r, column=1, value=lvl)
        cc.fill, cc.border = RISK_FILL[lvl], BORDER
        if lvl == "高风险":
            cc.font = HIGH_FONT
        dc = ws.cell(row=r, column=2, value=risk_desc[lvl])
        dc.border, dc.alignment = BORDER, WRAP_TOP
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="数据源状态图例 Status Legend").font = Font(bold=True, color=NAVY)
    r += 1
    for v in STATUS_LABEL.values():
        ws.cell(row=r, column=1, value=v).border = BORDER
        r += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 70
    ws.freeze_panes = "A2"
    return ws


def risk_mix(reg, imp, regs, insp, sent, labor=None, building=None, environment=None, consumer=None):
    mods = {"food_safety": reg, "import": imp, "regulation": regs,
            "labor": labor or [], "building": building or [], "environment": environment or [],
            "consumer": consumer or [], "inspection": insp, "sentiment": sent}
    return {m: {lvl: sum(1 for r in rows if r.get("riskLevel") == lvl) for lvl in RISK_LEVELS}
            for m, rows in mods.items()}


# ── bespoke sheet: 适用性矩阵 Applicability Matrix (from applicability_verdicts.json) ──
APPLIES_LABEL = {
    "applies": "适用 Applies", "approaching": "临近 Approaching", "not_yet": "暂不适用 Not yet",
    "always": "始终适用 Always", "na": "待补充/待核实 Pending",
}
APPLIES_FILL = {
    "applies": PatternFill("solid", fgColor="F4CCCC"), "approaching": PatternFill("solid", fgColor="FCE4D6"),
    "not_yet": PatternFill("solid", fgColor="E2EFDA"), "always": PatternFill("solid", fgColor="DDEBF7"),
    "na": PatternFill("solid", fgColor="E7EBF0"),
}


def build_applicability_sheet(wb, verdicts):
    ws = wb.create_sheet("适用性矩阵")
    headers = ["规则\nRule", "领域\nDomain", "辖区\nJurisdiction", "阈值\nThreshold",
               "我方当前值\nOur Value", "适用?\nApplies?", "距阈值\nDistance", "生效\nEffective", "依据链接\nSource"]
    style_header_row(ws, headers)
    for i, v in enumerate(verdicts):
        row = 2 + i
        status = v.get("status", "na")
        verdict_txt = "待核实 To verify" if (status == "na" and v.get("needsVerification")) else APPLIES_LABEL.get(status, status)
        our = v.get("ourValue")
        our_txt = "待补充 Pending" if our is None and status == "na" else ("—" if our is None else our)
        dist = v.get("distance")
        vals = [v.get("nameZh") or v.get("nameEn", ""), MODULE_LABEL.get(v.get("module"), v.get("module", "")),
                v.get("jurisdiction", ""), v.get("threshold") if v.get("threshold") is not None else "—",
                our_txt, verdict_txt, ("—" if dist is None else dist),
                (v.get("effectiveDate") or "").replace("-", "/"), v.get("thresholdSourceUrl", "")]
        for ci, val in enumerate(vals):
            c = ws.cell(row=row, column=ci + 1, value=val)
            c.border, c.alignment = BORDER, WRAP_TOP
            if ci == 5 and status in APPLIES_FILL:  # the 适用? verdict cell
                c.fill = APPLIES_FILL[status]
                if status == "applies":
                    c.font = HIGH_FONT
    for ci, w in enumerate([34, 22, 16, 12, 14, 18, 12, 13, 44]):
        ws.column_dimensions[get_column_letter(ci + 1)].width = w
    finalize_table(ws, len(headers), len(verdicts))
    return ws


def verify(path, expect_high_fill):
    wb = load_workbook(path)
    titles = wb.sheetnames
    expect = ["月度摘要", "食品安全主表", "进口出口监管", "州地方法规", "用工合规", "建筑与职业安全",
              "环境卫生", "消费者与员工保护", "咖啡馆检查", "适用性矩阵", "数据源日志", "字段说明"]
    assert titles == expect, f"sheet titles/order: {titles}"
    hdr5 = [c.value for c in wb["咖啡馆检查"][1]]
    assert any(h and "门店编号" in str(h) for h in hdr5), "missing 门店编号 col on 咖啡馆检查"
    # Applicability sheet must carry the 适用? verdict column + at least one rule row.
    hdr_ap = [c.value for c in wb["适用性矩阵"][1]]
    assert any(h and "适用?" in str(h) for h in hdr_ap), "missing 适用? col on 适用性矩阵"
    assert wb["适用性矩阵"].max_row >= 2, "no rule rows on 适用性矩阵"
    # The high-risk fill is a property of the DATA — only assert it when a 高风险 food-safety row exists
    # (a legitimate low-risk month must not fail the export).
    if expect_high_fill:
        found_fill = any(
            c.fill and c.fill.fgColor and c.fill.fgColor.rgb and str(c.fill.fgColor.rgb).endswith("F4CCCC")
            for row in wb["食品安全主表"].iter_rows(min_row=2) for c in row
        )
        assert found_fill, "no high-risk fill found on 食品安全主表 despite 高风险 rows present"
    assert wb["数据源日志"].max_row >= 2, "no pull-log rows on 数据源日志"
    print(f"verify OK: {len(titles)} sheets — {titles}")


def main():
    reg = [r for r in load("regulatory.json") if servable(r)]
    insp = [r for r in load("inspections.json") if servable(r)]
    imp = [r for r in load("import_export.json") if servable(r)]
    regs = [r for r in load("regulations.json") if servable(r)]
    sent = [r for r in load("sentiment.json") if servable(r) and not r.get("excluded")]
    labor = [r for r in load("labor.json") if servable(r)]
    building = [r for r in load("building.json") if servable(r)]
    environment = [r for r in load("environment.json") if servable(r)]
    consumer = [r for r in load("consumer.json") if servable(r)]
    verdicts = load("applicability_verdicts.json")
    meta = load("meta.json")

    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    build_summary_sheet(wb, meta, risk_mix(reg, imp, regs, insp, sent, labor, building, environment, consumer), verdicts)
    build_table_sheet(wb, "食品安全主表", SHEET1_COLS, HDR1, reg)
    build_table_sheet(wb, "进口出口监管", SHEET3_COLS, HDR3, imp)
    build_table_sheet(wb, "州地方法规", SHEET4_COLS, HDR4, regs)
    build_table_sheet(wb, "用工合规", SHEET5_COLS, HDR5, labor)
    build_table_sheet(wb, "建筑与职业安全", SHEET6_COLS, HDR6, building)
    build_table_sheet(wb, "环境卫生", SHEET7_COLS, HDR7, environment)
    build_table_sheet(wb, "消费者与员工保护", SHEET8_COLS, HDR8, consumer)
    build_table_sheet(wb, "咖啡馆检查", SHEET2_COLS, HDR2, insp)
    build_applicability_sheet(wb, verdicts)
    build_sources_sheet(wb, meta)
    build_field_guide_sheet(wb)

    os.makedirs(OUT_DIR, exist_ok=True)
    wb.save(OUT)
    print(f"export: 12 sheets — summary · {len(reg)} food-safety · {len(imp)} import · {len(regs)} regulation · "
          f"{len(labor)} labor · {len(building)} building · {len(environment)} env · {len(consumer)} consumer · "
          f"{len(insp)} inspection · {len(verdicts)} applicability · {len(meta.get('provenance', []))} sources "
          f"→ public/exports/monthly_report.xlsx")
    verify(OUT, expect_high_fill=any(r.get("riskLevel") == "高风险" for r in reg))


if __name__ == "__main__":
    main()
